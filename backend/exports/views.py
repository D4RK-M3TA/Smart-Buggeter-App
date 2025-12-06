import csv
import io
from datetime import date
from django.http import HttpResponse
from rest_framework import views, status
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from drf_spectacular.utils import extend_schema, OpenApiParameter
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

from transactions.models import Transaction, Category
from budgets.models import Budget


class ExportTransactionsCSVView(views.APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=['Exports'],
        parameters=[
            OpenApiParameter('start_date', str, description='Start date (YYYY-MM-DD)'),
            OpenApiParameter('end_date', str, description='End date (YYYY-MM-DD)'),
            OpenApiParameter('category', str, description='Category UUID'),
        ]
    )
    def get(self, request):
        queryset = Transaction.objects.filter(user=request.user)
        
        if request.query_params.get('start_date'):
            queryset = queryset.filter(date__gte=request.query_params['start_date'])
        if request.query_params.get('end_date'):
            queryset = queryset.filter(date__lte=request.query_params['end_date'])
        if request.query_params.get('category'):
            queryset = queryset.filter(category_id=request.query_params['category'])
        
        queryset = queryset.select_related('category').order_by('-date')
        
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="transactions_{date.today()}.csv"'
        
        writer = csv.writer(response)
        writer.writerow(['Date', 'Description', 'Amount', 'Type', 'Category', 'Notes', 'Is Recurring'])
        
        for tx in queryset:
            writer.writerow([
                tx.date.isoformat(),
                tx.description,
                str(tx.amount),
                tx.transaction_type,
                tx.category.name if tx.category else '',
                tx.notes,
                'Yes' if tx.is_recurring else 'No'
            ])
        
        return response


class ExportTransactionsExcelView(views.APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=['Exports'],
        parameters=[
            OpenApiParameter('start_date', str, description='Start date (YYYY-MM-DD)'),
            OpenApiParameter('end_date', str, description='End date (YYYY-MM-DD)'),
            OpenApiParameter('category', str, description='Category UUID'),
        ]
    )
    def get(self, request):
        queryset = Transaction.objects.filter(user=request.user)
        
        if request.query_params.get('start_date'):
            queryset = queryset.filter(date__gte=request.query_params['start_date'])
        if request.query_params.get('end_date'):
            queryset = queryset.filter(date__lte=request.query_params['end_date'])
        if request.query_params.get('category'):
            queryset = queryset.filter(category_id=request.query_params['category'])
        
        queryset = queryset.select_related('category').order_by('-date')
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Transactions"
        
        headers = ['Date', 'Description', 'Amount', 'Type', 'Category', 'Notes', 'Is Recurring']
        ws.append(headers)
        
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        
        for tx in queryset:
            ws.append([
                tx.date.isoformat(),
                tx.description,
                float(tx.amount),
                tx.transaction_type,
                tx.category.name if tx.category else '',
                tx.notes,
                'Yes' if tx.is_recurring else 'No'
            ])
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="transactions_{date.today()}.xlsx"'
        
        return response


class ExportBudgetReportPDFView(views.APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=['Exports'],
        parameters=[
            OpenApiParameter('year', int, description='Year'),
            OpenApiParameter('month', int, description='Month (1-12)'),
        ]
    )
    def get(self, request):
        from django.db.models import Sum
        from calendar import monthrange
        
        today = date.today()
        year = int(request.query_params.get('year', today.year))
        month = int(request.query_params.get('month', today.month))
        
        start_date = date(year, month, 1)
        _, last_day = monthrange(year, month)
        end_date = date(year, month, last_day)
        
        transactions = Transaction.objects.filter(
            user=request.user,
            date__gte=start_date,
            date__lte=end_date
        )
        
        income = transactions.filter(
            transaction_type='credit'
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        expenses = transactions.filter(
            transaction_type='debit'
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        by_category = transactions.filter(
            transaction_type='debit',
            category__isnull=False
        ).values('category__name').annotate(
            total=Sum('amount')
        ).order_by('-total')
        
        budgets = Budget.objects.filter(
            user=request.user,
            start_date__lte=end_date,
            end_date__gte=start_date
        ).select_related('category')
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=0.5*inch)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=20
        )
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=14,
            spaceAfter=10,
            spaceBefore=15
        )
        
        month_name = start_date.strftime('%B %Y')
        elements.append(Paragraph(f"Budget Report - {month_name}", title_style))
        
        elements.append(Paragraph("Summary", heading_style))
        summary_data = [
            ['Metric', 'Amount'],
            ['Total Income', f'${income:,.2f}'],
            ['Total Expenses', f'${expenses:,.2f}'],
            ['Net', f'${(income - expenses):,.2f}'],
        ]
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(summary_table)
        
        if by_category:
            elements.append(Paragraph("Spending by Category", heading_style))
            category_data = [['Category', 'Amount', '% of Total']]
            for cat in by_category:
                pct = (cat['total'] / expenses * 100) if expenses > 0 else 0
                category_data.append([
                    cat['category__name'],
                    f"${cat['total']:,.2f}",
                    f"{pct:.1f}%"
                ])
            
            cat_table = Table(category_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
            cat_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            elements.append(cat_table)
        
        if budgets:
            elements.append(Paragraph("Budget Status", heading_style))
            budget_data = [['Budget', 'Limit', 'Spent', 'Remaining', 'Status']]
            for budget in budgets:
                pct = budget.percentage_used
                status_text = 'On Track' if pct < 80 else ('Warning' if pct < 100 else 'Exceeded')
                budget_data.append([
                    budget.name,
                    f"${budget.amount:,.2f}",
                    f"${budget.spent:,.2f}",
                    f"${budget.remaining:,.2f}",
                    status_text
                ])
            
            budget_table = Table(budget_data, colWidths=[1.8*inch, 1.2*inch, 1.2*inch, 1.2*inch, 1*inch])
            budget_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ]))
            elements.append(budget_table)
        
        elements.append(Spacer(1, 20))
        elements.append(Paragraph(
            f"Generated on {date.today().strftime('%B %d, %Y')}",
            styles['Normal']
        ))
        
        doc.build(elements)
        
        buffer.seek(0)
        response = HttpResponse(buffer.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="budget_report_{year}_{month:02d}.pdf"'
        
        return response


class ExportCategorySummaryView(views.APIView):
    permission_classes = [IsAuthenticated]

    @extend_schema(
        tags=['Exports'],
        parameters=[
            OpenApiParameter('format', str, description='Export format: csv or excel'),
            OpenApiParameter('start_date', str, description='Start date (YYYY-MM-DD)'),
            OpenApiParameter('end_date', str, description='End date (YYYY-MM-DD)'),
        ]
    )
    def get(self, request):
        from django.db.models import Sum, Count
        
        export_format = request.query_params.get('format', 'csv')
        
        queryset = Transaction.objects.filter(
            user=request.user,
            transaction_type='debit',
            category__isnull=False
        )
        
        if request.query_params.get('start_date'):
            queryset = queryset.filter(date__gte=request.query_params['start_date'])
        if request.query_params.get('end_date'):
            queryset = queryset.filter(date__lte=request.query_params['end_date'])
        
        summary = queryset.values(
            'category__name'
        ).annotate(
            total=Sum('amount'),
            count=Count('id')
        ).order_by('-total')
        
        if export_format == 'excel':
            wb = Workbook()
            ws = wb.active
            ws.title = "Category Summary"
            
            ws.append(['Category', 'Total Amount', 'Transaction Count'])
            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True)
            
            for row in summary:
                ws.append([row['category__name'], float(row['total']), row['count']])
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="category_summary_{date.today()}.xlsx"'
        else:
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="category_summary_{date.today()}.csv"'
            
            writer = csv.writer(response)
            writer.writerow(['Category', 'Total Amount', 'Transaction Count'])
            
            for row in summary:
                writer.writerow([row['category__name'], str(row['total']), row['count']])
        
        return response
